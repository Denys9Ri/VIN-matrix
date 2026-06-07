import csv
import io
import json
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .crm_client_views import build_clients
from .models import InventoryItem, OrderPart, OrderService, Visit
from .safe_crm_views import safe_ensure_company


def _money(value):
    try:
        return float(Decimal(str(value or 0)))
    except Exception:
        return 0.0


def _csv_response(filename, headers, rows):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response, delimiter=';')
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def _date(value):
    if not value:
        return ''
    try:
        return timezone.localtime(value).strftime('%d.%m.%Y %H:%M')
    except Exception:
        return str(value)


def _visit_totals(visit):
    parts_total = sum(_money(p.sell_price) * (_money(p.quantity) or 1) for p in visit.parts.all())
    parts_cost = sum(_money(p.buy_price) * (_money(p.quantity) or 1) for p in visit.parts.all())
    services_total = sum(_money(s.price) * (_money(s.quantity) or 1) for s in visit.services.all())
    paid = _money(getattr(visit, 'prepayment_amount', 0))
    total = parts_total + services_total
    return {
        'parts_total': round(parts_total, 2),
        'services_total': round(services_total, 2),
        'total': round(total, 2),
        'cost': round(parts_cost, 2),
        'profit': round(total - parts_cost, 2),
        'paid': round(paid, 2),
        'debt': round(max(total - paid, 0), 2),
    }


class ClientsExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = safe_ensure_company(request.user)
        if not company:
            return Response({'error': 'Немає компанії'}, status=403)
        clients = build_clients(company, search=request.query_params.get('search', '').strip())
        rows = []
        for c in clients:
            rows.append([
                c.get('client') or 'Без імені',
                c.get('phone') or '',
                c.get('orders_count') or 0,
                c.get('total_revenue') or 0,
                c.get('paid_amount') or 0,
                c.get('debt_amount') or 0,
                c.get('total_profit') or 0,
                c.get('status') or '',
                _date(c.get('last_order_date')),
                ', '.join([x.get('plate') or '' for x in c.get('cars', []) if x.get('plate')]),
            ])
        return _csv_response('vin_matrix_clients.csv', ['Клієнт', 'Телефон', 'Замовлень/візитів', 'Сума продажів', 'Оплачено', 'Борг', 'Прибуток', 'Статус', 'Остання дата', 'Авто'], rows)


class OrdersExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = safe_ensure_company(request.user)
        if not company:
            return Response({'error': 'Немає компанії'}, status=403)
        qs = Visit.objects.filter(company=company).prefetch_related('parts', 'services').order_by('-created_at')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        rows = []
        for v in qs:
            totals = _visit_totals(v)
            rows.append([
                v.id,
                'Магазин' if getattr(company, 'business_type', '') == 'store' else 'СТО',
                v.status,
                v.client,
                v.phone,
                v.plate,
                v.vin_code,
                _date(v.scheduled_datetime or v.created_at),
                totals['parts_total'],
                totals['services_total'],
                totals['total'],
                totals['paid'],
                totals['debt'],
                totals['profit'],
                v.delivery_type or '',
                v.delivery_data or '',
                v.comment or '',
            ])
        return _csv_response('vin_matrix_orders.csv', ['ID', 'Режим', 'Статус', 'Клієнт', 'Телефон', 'Авто/номер', 'VIN', 'Дата', 'Товари', 'Роботи', 'Разом', 'Оплачено', 'Борг', 'Прибуток', 'Доставка', 'Дані доставки/ТТН', 'Коментар'], rows)


class InventoryExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = safe_ensure_company(request.user)
        if not company:
            return Response({'error': 'Немає компанії'}, status=403)
        qs = InventoryItem.objects.filter(company=company).select_related('category', 'supplier').order_by('brand', 'article')
        rows = []
        for item in qs:
            qty = int(item.quantity or 0)
            buy = _money(item.buy_price)
            sell = _money(item.sell_price)
            rows.append([
                item.brand,
                item.article,
                item.name,
                qty,
                buy,
                sell,
                round(qty * buy, 2),
                round(qty * sell, 2),
                item.category.name if item.category else '',
                item.supplier.name if item.supplier else '',
                _date(item.updated_at),
            ])
        return _csv_response('vin_matrix_inventory.csv', ['Бренд', 'Артикул', 'Назва', 'Кількість', 'Закупка', 'Продаж', 'Сума закупки', 'Сума продажу', 'Категорія', 'Постачальник', 'Оновлено'], rows)


class BackupExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company = safe_ensure_company(request.user)
        if not company:
            return Response({'error': 'Немає компанії'}, status=403)
        visits = Visit.objects.filter(company=company).prefetch_related('parts', 'services').order_by('id')
        inventory = InventoryItem.objects.filter(company=company).select_related('category', 'supplier').order_by('id')
        payload = {
            'exported_at': timezone.now().isoformat(),
            'company': {'id': company.id, 'name': company.name, 'business_type': company.business_type},
            'clients': build_clients(company),
            'orders': [
                {
                    'id': v.id,
                    'client': v.client,
                    'phone': v.phone,
                    'plate': v.plate,
                    'vin_code': v.vin_code,
                    'status': v.status,
                    'delivery_type': v.delivery_type,
                    'delivery_data': v.delivery_data,
                    'payment_status': v.payment_status,
                    'prepayment_amount': str(v.prepayment_amount or 0),
                    'created_at': v.created_at.isoformat() if v.created_at else None,
                    'scheduled_datetime': v.scheduled_datetime.isoformat() if v.scheduled_datetime else None,
                    'comment': v.comment,
                    'parts': [{'brand': p.brand, 'article': p.article, 'name': p.name, 'quantity': str(p.quantity), 'buy_price': str(p.buy_price), 'sell_price': str(p.sell_price), 'supplier': p.supplier, 'status': p.status} for p in v.parts.all()],
                    'services': [{'name': s.name, 'price': str(s.price), 'quantity': str(s.quantity), 'status': s.status} for s in v.services.all()],
                }
                for v in visits
            ],
            'inventory': [
                {'brand': i.brand, 'article': i.article, 'name': i.name, 'quantity': i.quantity, 'buy_price': str(i.buy_price), 'sell_price': str(i.sell_price), 'category': i.category.name if i.category else '', 'supplier': i.supplier.name if i.supplier else ''}
                for i in inventory
            ],
        }
        response = JsonResponse(payload, json_dumps_params={'ensure_ascii': False, 'indent': 2})
        response['Content-Disposition'] = 'attachment; filename="vin_matrix_backup.json"'
        return response


class LegacyClientsImportView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        company = safe_ensure_company(request.user)
        if not company:
            return Response({'error': 'Немає компанії'}, status=403)
        upload = request.FILES.get('file')
        if not upload:
            return Response({'error': 'Додайте CSV або XLSX файл з клієнтами.'}, status=400)
        rows = _read_upload(upload)
        if not rows:
            return Response({'error': 'Файл порожній або не читається.'}, status=400)
        headers = [_norm(h) for h in rows[0]]
        aliases = {
            'client': ['клієнт', 'піб', 'імя', 'імʼя', 'имя', 'name', 'client', 'fullname', 'full_name'],
            'phone': ['телефон', 'phone', 'mobile', 'номер'],
            'plate': ['авто', 'номер', 'держномер', 'plate', 'car'],
            'vin_code': ['vin', 'vinкод', 'vin_code'],
            'comment': ['коментар', 'comment', 'note', 'примітка'],
        }
        idx = {}
        for key, names in aliases.items():
            for name in names:
                if name in headers:
                    idx[key] = headers.index(name)
                    break
        if 'client' not in idx and 'phone' not in idx:
            return Response({'error': 'У файлі має бути хоча б колонка Клієнт або Телефон.'}, status=400)
        created = skipped = 0
        for row in rows[1:]:
            def cell(key):
                pos = idx.get(key)
                return str(row[pos] or '').strip() if pos is not None and pos < len(row) else ''
            client = cell('client') or 'Старий клієнт'
            phone = cell('phone')
            plate = cell('plate') or 'CLIENT'
            vin = cell('vin_code')[:17]
            comment = cell('comment')
            if not client and not phone:
                skipped += 1
                continue
            exists = Visit.objects.filter(company=company, phone=phone).exists() if phone else False
            if exists:
                skipped += 1
                continue
            Visit.objects.create(company=company, client=client, phone=phone, plate=plate, vin_code=vin, status='CLIENT', comment=f'Імпорт старої бази клієнтів. {comment}'.strip())
            created += 1
        return Response({'created': created, 'skipped': skipped, 'message': f'Імпортовано клієнтів: {created}. Пропущено: {skipped}.'})


def _norm(value):
    return str(value or '').strip().lower().replace(' ', '').replace('_', '').replace('-', '')


def _read_upload(upload):
    name = (upload.name or '').lower()
    if name.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(upload, read_only=True, data_only=True)
            return list(wb.active.iter_rows(values_only=True))
        except Exception:
            return []
    try:
        raw = upload.read().decode('utf-8-sig')
    except Exception:
        upload.seek(0)
        raw = upload.read().decode('cp1251', errors='ignore')
    sample = raw[:2048]
    delimiter = ';' if sample.count(';') >= sample.count(',') else ','
    return list(csv.reader(io.StringIO(raw), delimiter=delimiter))
