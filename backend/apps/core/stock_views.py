from decimal import Decimal
import re

from django.db import transaction
from django.db.models import Q
from rest_framework import serializers, viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Category, Company, InventoryItem, OrderPart, StockMovement, Supplier
from .safe_crm_views import safe_ensure_company


def to_decimal(value):
    try:
        return Decimal(str(value or 0).replace(',', '.').replace(' ', ''))
    except Exception:
        return Decimal('0')


def to_int(value):
    try:
        return max(int(float(str(value or 1).replace(',', '.').replace(' ', ''))), 1)
    except Exception:
        return 1


def apply_company_margin(buy_price, sell_price, company):
    buy = to_decimal(buy_price)
    sell = to_decimal(sell_price)
    if sell > 0 or buy <= 0:
        return sell
    try:
        margin = Decimal(str(getattr(company, 'global_margin_percent', 0) or 0))
    except Exception:
        margin = Decimal('0')
    calculated = buy + (buy * margin / Decimal('100'))
    return calculated.quantize(Decimal('0.01'))


def norm(value):
    return str(value or '').strip().lower().replace(' ', '').replace('_', '')


def movement_display(obj):
    note = str(obj.note or '').lower()
    qty = int(obj.quantity or 0)
    if 'резерв' in note and qty == 0:
        return 'Резерв'
    if 'зняття резерв' in note or 'резерв знято' in note:
        return 'Зняття резерву'
    if 'повернення' in note and qty > 0:
        return 'Повернення на склад'
    if 'брак' in note:
        return 'Брак / списання'
    if qty < 0:
        return 'Списання / продаж'
    if obj.movement_type == 'receipt' or qty > 0:
        return 'Прихід'
    return 'Ручне коригування'


def movement_order_id(obj):
    if obj.source_order_part_id:
        return obj.source_order_part.visit_id if obj.source_order_part else None
    match = re.search(r'№\s*(\d+)', str(obj.note or ''))
    return match.group(1) if match else None


class StockMovementSerializer(serializers.ModelSerializer):
    supplier_name = serializers.CharField(source='supplier.name', read_only=True)
    total_sum = serializers.SerializerMethodField()
    movement_label = serializers.SerializerMethodField()
    quantity_display = serializers.SerializerMethodField()
    order_id = serializers.SerializerMethodField()
    reason_display = serializers.SerializerMethodField()

    class Meta:
        model = StockMovement
        fields = '__all__'
        read_only_fields = ['company', 'inventory_item', 'created_by', 'created_at']

    def get_total_sum(self, obj):
        return round(float(obj.buy_price or 0) * abs(float(obj.quantity or 0)), 2)

    def get_movement_label(self, obj):
        return movement_display(obj)

    def get_quantity_display(self, obj):
        qty = int(obj.quantity or 0)
        if qty > 0:
            return f'+{qty}'
        if qty < 0:
            return str(qty)
        return '0'

    def get_order_id(self, obj):
        return movement_order_id(obj)

    def get_reason_display(self, obj):
        note = str(obj.note or '').strip()
        if not note:
            return ''
        return note


class StockMovementViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        company = safe_ensure_company(self.request.user)
        qs = StockMovement.objects.filter(company=company) if company else StockMovement.objects.none()
        search = self.request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(Q(brand__icontains=search) | Q(article__icontains=search) | Q(name__icontains=search) | Q(supplier__name__icontains=search) | Q(note__icontains=search))
        return qs.select_related('supplier', 'inventory_item', 'created_by', 'source_order_part', 'source_order_part__visit')


class StockReceiveViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated]

    def _receive_payload(self, request, data):
        company = safe_ensure_company(request.user)
        if not company:
            return None, Response({'error': 'Немає компанії'}, status=400)

        brand = str(data.get('brand') or '').strip().upper()
        article = str(data.get('article') or '').strip().upper()
        name = str(data.get('name') or '').strip()
        if not brand or not article or not name:
            return None, Response({'error': 'Бренд, артикул і назва обовʼязкові'}, status=400)

        quantity = to_int(data.get('quantity'))
        buy_price = to_decimal(data.get('buy_price'))

        # Lock the company row first. InventoryItem has no uniqueness constraint on
        # company + brand + article, so this prevents simultaneous receipts from
        # creating duplicate rows for the same item.
        with transaction.atomic():
            company = Company.objects.select_for_update().get(pk=company.pk)
            sell_price = apply_company_margin(buy_price, data.get('sell_price'), company)

            supplier = None
            if data.get('supplier'):
                supplier = Supplier.objects.filter(company=company, id=data.get('supplier')).first()
            elif data.get('supplier_name'):
                supplier, _ = Supplier.objects.get_or_create(company=company, name=str(data.get('supplier_name')).strip())

            category = Category.objects.filter(company=company, id=data.get('category')).first() if data.get('category') else None
            if not category and data.get('category_name'):
                category, _ = Category.objects.get_or_create(company=company, name=str(data.get('category_name')).strip())

            order_part = OrderPart.objects.filter(id=data.get('order_part_id'), visit__company=company).first() if data.get('order_part_id') else None
            item = InventoryItem.objects.select_for_update().filter(company=company, brand__iexact=brand, article__iexact=article).first()
            created = False
            if item:
                item.quantity = int(item.quantity or 0) + quantity
                item.name = name
                item.buy_price = buy_price
                item.sell_price = sell_price or item.sell_price
                item.category = category or item.category
                item.supplier = supplier or item.supplier
                item.save()
            else:
                created = True
                item = InventoryItem.objects.create(company=company, category=category, supplier=supplier, brand=brand, article=article, name=name, quantity=quantity, buy_price=buy_price, sell_price=sell_price)

            movement = StockMovement.objects.create(company=company, inventory_item=item, supplier=supplier, source_order_part=order_part, brand=brand, article=article, name=name, quantity=quantity, buy_price=buy_price, sell_price=sell_price, note=data.get('note') or 'Прихід товару', created_by=request.user)
            if order_part:
                order_part.status = 'ARRIVED'
                order_part.save(update_fields=['status'])

            result = {
                'created': created,
                'item_id': item.id,
                'quantity': item.quantity,
                'sell_price': sell_price,
                'movement': StockMovementSerializer(movement).data,
            }
        return result, None

    @action(detail=False, methods=['post'], url_path='receive')
    def receive(self, request):
        result, error = self._receive_payload(request, request.data or {})
        if error:
            return error
        return Response(result, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'], url_path='import-file')
    def import_file(self, request):
        upload = request.FILES.get('file') or request.FILES.get('price_file')
        if not upload:
            return Response({'error': 'Додайте Excel-файл .xlsx'}, status=400)
        try:
            from openpyxl import load_workbook
        except Exception:
            return Response({'error': 'На сервері не встановлено openpyxl для читання Excel.'}, status=500)
        try:
            wb = load_workbook(upload, read_only=True, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
        except Exception:
            return Response({'error': 'Не вдалося прочитати Excel-файл.'}, status=400)
        if not rows:
            return Response({'error': 'Файл порожній.'}, status=400)

        headers = [norm(x) for x in rows[0]]
        aliases = {
            'brand': ['бренд', 'brand', 'виробник', 'manufacturer'],
            'article': ['артикул', 'код', 'article', 'partnumber', 'номер', 'sku'],
            'name': ['назва', 'найменування', 'опис', 'name', 'description'],
            'quantity': ['кількість', 'ксть', 'залишок', 'quantity', 'qty', 'stock'],
            'buy_price': ['закупка', 'закупочна', 'ціназакупки', 'buyprice', 'purchaseprice', 'cost'],
            'sell_price': ['продаж', 'цінапродажу', 'sellprice', 'price'],
            'supplier_name': ['постачальник', 'supplier'],
            'category_name': ['категорія', 'category'],
        }
        idx = {}
        for key, names in aliases.items():
            for name in names:
                if name in headers:
                    idx[key] = headers.index(name)
                    break
        required = ['brand', 'article', 'name']
        missing = [x for x in required if x not in idx]
        if missing:
            return Response({'error': 'У файлі мають бути колонки: бренд, артикул, назва.'}, status=400)

        created = updated = skipped = 0
        try:
            # The whole file is atomic: an unexpected database failure cannot leave
            # a partially imported inventory behind.
            with transaction.atomic():
                for row in rows[1:]:
                    def cell(key, default=''):
                        pos = idx.get(key)
                        return row[pos] if pos is not None and pos < len(row) else default

                    data = {
                        'brand': cell('brand'),
                        'article': cell('article'),
                        'name': cell('name'),
                        'quantity': cell('quantity', 1),
                        'buy_price': cell('buy_price', 0),
                        'sell_price': cell('sell_price', 0),
                        'supplier_name': cell('supplier_name', request.data.get('supplier_name') or ''),
                        'category_name': cell('category_name', ''),
                        'note': 'Імпорт з Excel',
                    }
                    if not str(data['brand'] or '').strip() or not str(data['article'] or '').strip() or not str(data['name'] or '').strip():
                        skipped += 1
                        continue
                    result, error = self._receive_payload(request, data)
                    if error:
                        skipped += 1
                        continue
                    if result.get('created'):
                        created += 1
                    else:
                        updated += 1
        except Exception:
            return Response({'error': 'Імпорт не завершено. Дані не були частково збережені.'}, status=500)

        return Response({'created': created, 'updated': updated, 'skipped': skipped})
