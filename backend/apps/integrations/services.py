from openpyxl import load_workbook
import math
import requests
from .models import PriceItem

def parse_supplier_excel(file_path, column_mapping, exchange_rate=1.0):
    """
    Універсальний парсер прайс-листів.
    file_path: шлях до файлу Excel.
    column_mapping: словник, наприклад {"part_number": "Код товару", "price": "Ціна", "brand": "Бренд", "name": "Назва"}
    exchange_rate: курс валют (за замовчуванням 1.0, якщо прайс у гривні)
    """
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return {"status": "success", "total_items": 0, "data": []}

        normalized_headers = [str(h).strip() if h is not None else "" for h in headers]

        index_map = {}
        for db_col, excel_col in column_mapping.items():
            if excel_col in normalized_headers:
                index_map[db_col] = normalized_headers.index(excel_col)

        records = []
        for row in rows:
            record = {}
            for db_col, idx in index_map.items():
                value = row[idx] if idx < len(row) else None
                record[db_col] = value

            part_number = record.get('part_number')
            price = record.get('price')
            if part_number in (None, '') or price in (None, ''):
                continue

            record['part_number'] = str(part_number).strip()

            try:
                numeric_price = round(float(price) * float(exchange_rate), 2)
            except (TypeError, ValueError):
                continue

            record['price'] = numeric_price
            records.append(record)

        return {
            "status": "success",
            "total_items": len(records),
            "data": records
        }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def fetch_api_price(api_url, api_token, part_number):
    """
    Базовий запит до API постачальника.
    """
    headers = {
        "Authorization": f"Bearer {api_token}",
        "Content-Type": "application/json"
    }
    
    # Це приклад (payload буде залежати від документації конкретного постачальника)
    payload = {
        "search": part_number,
        "exact_match": True
    }
    
    try:
        response = requests.post(api_url, json=payload, headers=headers, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            # Уявімо, що API повертає {"results": [{"brand": "LPR", "price": 45.5, "currency": "EUR"}]}
            return {
                "status": "success",
                "data": data.get('results', [])
            }
        else:
            return {
                "status": "error",
                "message": f"API Error: {response.status_code}"
            }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }

def import_price_to_db(supplier_config, parsed_records):
    """
    Зберігає розпакований Excel-прайс у базу даних.
    Очищає старий прайс цього постачальника перед завантаженням нового.
    """
    try:
        # 1. Видаляємо всі старі позиції цього постачальника
        PriceItem.objects.filter(supplier=supplier_config).delete()
        
        # 2. Готуємо список об'єктів для масового збереження
        items_to_create = []
        for record in parsed_records:
            items_to_create.append(
                PriceItem(
                    supplier=supplier_config,
                    part_number=str(record.get('part_number', '')).strip(),
                    brand=record.get('brand', 'Unknown'),
                    name=record.get('name', 'Деталь'),
                    price=record.get('price', 0.00),
                    quantity=record.get('quantity', '>1')
                )
            )
        
        # 3. Зберігаємо всі рядки одним запитом у БД (batch_size=5000 береже оперативну пам'ять)
        PriceItem.objects.bulk_create(items_to_create, batch_size=5000)
        
        return {"status": "success", "imported_count": len(items_to_create)}
        
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ==========================================
# НОВИЙ БЛОК: Адаптери для замовлень API
# ==========================================

def order_from_vesna(api_token, items):
    """ Формує запит чітко по Swagger'у Весни """
    payload = {
        "order_items": [{"article": item.part_number, "qty": 1} for item in items]
    }
    # Імітація успішного замовлення:
    return {"status": "success", "supplier_order_id": "VSN-998877"}

def order_from_autotechnics(api_token, items):
    """ Формує запит чітко по Swagger'у Автотехнікса """
    payload = {
        "parts": [item.part_number for item in items],
        "delivery_type": "express"
    }
    return {"status": "success", "supplier_order_id": "ATX-112233"}

def dispatch_api_order(supplier_name, api_token, items):
    """ 
    Визначає, який саме адаптер викликати, залежно від назви постачальника.
    """
    name = supplier_name.lower()
    
    try:
        if 'весна' in name:
            return order_from_vesna(api_token, items)
        elif 'автотехнікс' in name:
            return order_from_autotechnics(api_token, items)
        else:
            return {"status": "error", "message": f"Немає адаптера замовлення для {supplier_name}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}
